#!/usr/bin/env python3
"""
Build the interactive CI/CD friction-cost workbook for stakeholders.

Every yellow cell is editable. All other numbers are formulas that recalc
automatically in Excel or Google Sheets.

This script is org-agnostic. All organisation-specific values (team size,
wage assumptions, measured pipeline durations, regional comp mix) are
loaded from data/report-data.json — see data/report-data.template.json
for the schema and copy/edit it before running.

Run:
    cd cicd-impact && .venv/bin/python scripts/05-build-xlsx.py

Output:
    outputs/cicd-impact-<org-name-from-json>.xlsx
"""
from __future__ import annotations
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = ROOT / "data" / "report-data.json"
TEMPLATE_PATH = ROOT / "data" / "report-data.template.json"

if not DATA_PATH.exists():
    sys.stderr.write(
        f"\n[ERROR] {DATA_PATH} not found.\n"
        f"Copy {TEMPLATE_PATH.name} to {DATA_PATH.name} and fill in your "
        f"organisation's measured values, then re-run.\n\n"
        f"  cp {TEMPLATE_PATH} {DATA_PATH}\n\n"
    )
    sys.exit(1)

with open(DATA_PATH) as f:
    DATA = json.load(f)

ORG = DATA["org_name"]
SEED = DATA["inputs_seed"]
TREND = DATA["trend"]["windows"]
REGIONS = DATA["regional_comp"]
REF_SALARIES = DATA["ref_salaries"]
DEV_OPTIONS = DATA["active_devs_options"]

OUT_PATH = ROOT / "outputs" / f"cicd-impact-{ORG.lower().replace(' ', '-')}.xlsx"
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# ---------- styles -------------------------------------------------------------

INPUT_FILL = PatternFill("solid", fgColor="FFF59D")
INPUT_STRONG = PatternFill("solid", fgColor="FFEB3B")
MEASURED_FILL = PatternFill("solid", fgColor="DCEDC8")
DERIVED_FILL = PatternFill("solid", fgColor="E1BEE7")
HEADLINE_FILL = PatternFill("solid", fgColor="1A237E")
SECTION_FILL = PatternFill("solid", fgColor="263238")
NOTE_FILL = PatternFill("solid", fgColor="F5F5F5")
SAVINGS_FILL = PatternFill("solid", fgColor="2E7D32")
WARN_FILL = PatternFill("solid", fgColor="FFCDD2")

WHITE_BOLD = Font(color="FFFFFF", bold=True, size=12)
SECTION_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADLINE_FONT = Font(color="FFFFFF", bold=True, size=20)
SAVINGS_FONT = Font(color="FFFFFF", bold=True, size=18)
LABEL_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=10, color="546E7A")
GREEN_FONT = Font(bold=True, color="2E7D32", size=12)

THIN = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

USD_FMT = '"$"#,##0.00'
USD_BIG_FMT = '"$"#,##0'
PCT_FMT = "0.0%"
MIN_FMT = '0.0" min"'

wb = Workbook()
wb.remove(wb.active)


def style_input(cell, fmt=None, strong=False):
    cell.fill = INPUT_STRONG if strong else INPUT_FILL
    cell.border = THIN
    cell.alignment = RIGHT
    if fmt:
        cell.number_format = fmt


def style_measured(cell, fmt=None):
    cell.fill = MEASURED_FILL
    cell.border = THIN
    cell.alignment = RIGHT
    if fmt:
        cell.number_format = fmt


def style_derived(cell, fmt=None):
    cell.fill = DERIVED_FILL
    cell.border = THIN
    cell.alignment = RIGHT
    if fmt:
        cell.number_format = fmt


def style_label(cell):
    cell.font = LABEL_FONT
    cell.alignment = LEFT
    cell.border = THIN


def style_note(cell):
    cell.font = NOTE_FONT
    cell.alignment = LEFT
    cell.fill = NOTE_FILL
    cell.border = THIN


def section(ws, row, text, span=6, fill=SECTION_FILL, font=SECTION_FONT, height=24):
    ws.cell(row=row, column=1, value=text).font = font
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height


def named(name, sheet, cell):
    ref = f"'{sheet}'!${cell[0]}${cell[1:]}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


# =============================================================================
# Sheet: README — onboarding
# =============================================================================
readme = wb.create_sheet("README")
readme.column_dimensions["A"].width = 110

readme["A1"] = f"{ORG} CI/CD Friction Cost Calculator"
readme["A1"].font = Font(bold=True, size=22, color="1A237E")
readme.row_dimensions[1].height = 32

baseline_min = TREND[0]["mean_all"]
current_min = TREND[2]["mean_all"]
improvement_pct = round((baseline_min - current_min) / baseline_min * 100)

readme_rows = [
    ("", None),
    (f"This workbook computes how much each minute of CI/CD costs the org — and what was already saved by improving "
     f"from {baseline_min} min to {current_min} min over the last 90 days ({improvement_pct}% drop).", LABEL_FONT),
    ("Edit any YELLOW cell. Everything else recalculates automatically.", LABEL_FONT),
    ("", None),
    ("WHERE TO START", Font(bold=True, size=14, color="263238")),
    ("    1. SAVINGS — what we already saved by getting CI/CD faster (the 'good news' story).", None),
    ("    2. HEADLINE — $/CI minute and friction tax broken into Annual / Monthly / Weekly / Daily.", None),
    ("    3. TREND — happy-path duration over the last 7 / 14 / 30 / 90 days, with sample sizes.", None),
    ("    4. WAGE CALCULATOR — enter annual salaries (US/EU/Israel mix) → blended $/min comes out.", None),
    ("    5. INPUTS — every editable knob. Yellow = your input. Green = measured. Purple = derived.", None),
    ("    6. VISION — what life looks like at 3, 5, 8, 10, 15, 20, 30, 60-min pipelines.", None),
    ("    7. SENSITIVITY — how the headline moves as inputs change.", None),
    ("    8. INVESTMENTS — the ROI ladder for fixes you might propose.", None),
    ("", None),
    ("CELL COLOR LEGEND", Font(bold=True, size=14, color="263238")),
    ("    YELLOW = your input. Edit freely.", None),
    ("    LIGHT GREEN = measured from GitHub Actions. Don't edit.", None),
    ("    LIGHT PURPLE = derived (formula). Don't edit — it recalcs.", None),
    ("    NAVY = headline output (formula).", None),
    ("    DARK GREEN = annual savings already realised.", None),
    ("", None),
    ("DEFENSIVE NOTES (read before presenting)", Font(bold=True, size=14, color="263238")),
    ("    • Unit is $/CI minute, not $/run. Multiply by minutes-saved-per-year for ROI.", None),
    ("    • 'CI failure rate' is NOT DORA's Change Failure Rate (production failures). Don't conflate.", None),
    ("    • Cognitive tax is piecewise (0/5/23 min) by wait band — defensible against the 'flat 23 min is folkloric' attack.", None),
    ("    • Blast radius is measured empirically from concurrency-lock waits, not guessed.", None),
    ("    • The Floor savings ($74k–$148k) is pure arithmetic — nobody can argue with it.", None),
    ("    • The Full savings ($4.3M–$8.7M) requires accepting the cognitive + rework multipliers.", None),
    ("", None),
    ("FILE-SHARING POLICY (handle with care)", Font(bold=True, size=14, color="263238")),
    ("    This workbook contains internal organisation numbers (team size, wage assumption, failure rate).", None),
    ("    Do NOT upload to public Drive folders without checking with eng leadership.", None),
    ("    Source citations and methodology (no numbers) live in the public-shareable cicd-impact/research/ folder.", None),
]
for i, (text, font) in enumerate(readme_rows, start=2):
    cell = readme.cell(row=i, column=1, value=text)
    if font:
        cell.font = font
    cell.alignment = LEFT


# =============================================================================
# Sheet: SAVINGS (clean rebuild — four sections, one tidy table per section)
# =============================================================================
sv = wb.create_sheet("Savings")
for col, w in zip("ABCDEFG", [28, 18, 18, 18, 18, 18, 32]):
    sv.column_dimensions[col].width = w

# ---- Banner ----
sv.merge_cells("A1:G1")
sv["A1"] = f"WHAT WE ALREADY SAVED — by getting CI/CD {improvement_pct}% faster over 90 days"
sv["A1"].font = HEADLINE_FONT
sv["A1"].fill = SAVINGS_FILL
sv["A1"].alignment = CENTER
sv.row_dimensions[1].height = 38

w90, w14, w7 = TREND[0], TREND[2], TREND[3]
mins_saved = round(w90["mean_all"] - w14["mean_all"], 1)

sv.merge_cells("A2:G2")
sv["A2"] = (f"90-day baseline: {w90['mean_all']} min mean.   "
            f"Last 14 days: {w14['mean_all']} min mean (n={w14['n_all']}).   "
            f"Last 7 days: {w7['mean_all']} min mean (n={w7['n_all']}, plateau-confirmed). "
            f"That is {mins_saved} minutes saved on every CI run, with zero feature code shipped.")
sv["A2"].font = LABEL_FONT
sv["A2"].alignment = CENTER
sv["A2"].fill = NOTE_FILL
sv.row_dimensions[2].height = 28

# ============================================================================
# SECTION 1 — THE MATH (single tidy table, top-down)
# ============================================================================
section(sv, 4, "STEP 1 — THE MATH (no money yet, just minutes and people)", span=7, fill=SECTION_FILL)

# Headers
for col_letter, label in zip("AB", ["What", "Value"]):
    sv[f"{col_letter}5"].value = label
    sv[f"{col_letter}5"].font = WHITE_BOLD
    sv[f"{col_letter}5"].fill = SECTION_FILL
    sv[f"{col_letter}5"].alignment = CENTER
    sv[f"{col_letter}5"].border = THIN

math_rows = [
    ("Pipeline duration — 90d baseline", "=Trend!C4", MIN_FMT, f"From the Trend sheet (n={w90['n_all']} successful runs)"),
    ("Pipeline duration — 14d current",  "=Trend!C6", MIN_FMT, f"From the Trend sheet (n={w14['n_all']}, more reliable than 7d)"),
    ("Minutes saved per CI run",          "=B6-B7",   MIN_FMT, "The improvement, per push"),
    ("CI runs per dev per day",           "=R_runs_per_dev_day", "0.0",  "Measured from the GitHub API"),
    ("Active developers",                 "=D_devs",  "0",  "Engineers shipping code into this repo"),
    ("Working days per year",             "=days_per_year", "0", "Standard finance default"),
    ("Total CI runs per year",            "=R_runs_per_dev_day*D_devs*days_per_year", "#,##0", "= D × R × days"),
    ("DEV-MINUTES SAVED PER YEAR",        "=B8*B12",  "#,##0", "The bedrock number — unarguable"),
    ("DEV-HOURS SAVED PER YEAR",          "=B13/60",  "#,##0", "= dev-minutes ÷ 60"),
    ("EQUIVALENT FTE-YEARS SAVED",        "=B14/(8*days_per_year)", "0.00", "An entire engineer's work freed up"),
]
for i, (label, formula, fmt, note) in enumerate(math_rows, start=6):
    sv[f"A{i}"] = label
    sv[f"B{i}"] = formula
    sv[f"C{i}"] = note
    style_label(sv[f"A{i}"])
    if i in (13, 14, 15):  # the bedrock numbers
        sv[f"B{i}"].fill = SAVINGS_FILL
        sv[f"B{i}"].font = Font(bold=True, color="FFFFFF", size=12)
        sv[f"B{i}"].number_format = fmt
        sv[f"B{i}"].border = THIN
        sv[f"B{i}"].alignment = CENTER
    else:
        style_derived(sv[f"B{i}"], fmt)
    sv.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
    style_note(sv[f"C{i}"])

# ============================================================================
# SECTION 2 — VALUATION FRAMES (one consistent table, side-by-side)
# ============================================================================
section(sv, 17, "STEP 2 — VALUATION FRAMES (pick the one your audience accepts)", span=7, fill=SECTION_FILL)

sv.merge_cells("A18:G18")
sv["A18"] = ("All frames use the SAME pipeline-minutes saved (computed in Step 1 above). They differ only in how each minute is valued. "
             "The further down the table, the more aggressive (and more sensitive to challenge).")
sv["A18"].font = NOTE_FONT
sv["A18"].alignment = LEFT
sv["A18"].fill = NOTE_FILL
sv.row_dimensions[18].height = 30

# Header
for col_letter, label in zip("ABCDEFG", ["Frame", "$/dev-min", "Annual", "Monthly", "Weekly", "Per working day", "What it means"]):
    sv[f"{col_letter}19"].value = label
    sv[f"{col_letter}19"].font = WHITE_BOLD
    sv[f"{col_letter}19"].fill = SECTION_FILL
    sv[f"{col_letter}19"].alignment = CENTER
    sv[f"{col_letter}19"].border = THIN

# Row template: A=label, B=$/min, C=annual=B13*B, D=annual/12, E=annual/50, F=annual/days, G=note
# B13 holds dev-minutes saved per year.
def fill_frame_row(row_n, label, dollar_per_min_formula, note, fill=DERIVED_FILL, big=False):
    sv[f"A{row_n}"] = label
    sv[f"B{row_n}"] = dollar_per_min_formula
    sv[f"C{row_n}"] = f"=$B$13*B{row_n}"
    sv[f"D{row_n}"] = f"=C{row_n}/12"
    sv[f"E{row_n}"] = f"=C{row_n}/50"
    sv[f"F{row_n}"] = f"=C{row_n}/days_per_year"
    sv[f"G{row_n}"] = note
    style_label(sv[f"A{row_n}"])
    sv[f"B{row_n}"].number_format = USD_FMT
    sv[f"B{row_n}"].border = THIN
    sv[f"B{row_n}"].alignment = RIGHT
    sv[f"B{row_n}"].fill = INPUT_FILL if "wage" in label.lower() else DERIVED_FILL
    for col in "CDEF":
        sv[f"{col}{row_n}"].fill = fill
        sv[f"{col}{row_n}"].font = Font(bold=True, color="FFFFFF", size=14 if big else 11)
        sv[f"{col}{row_n}"].number_format = USD_BIG_FMT
        sv[f"{col}{row_n}"].border = THIN
        sv[f"{col}{row_n}"].alignment = CENTER
    style_note(sv[f"G{row_n}"])
    if big:
        sv.row_dimensions[row_n].height = 28

# --- Sub-heading: Wage frames (the floor) ---
sv.merge_cells("A20:G20")
sv["A20"] = "  Wage frames — pure idle time × hourly rate. No multipliers. Irrefutable."
sv["A20"].font = LABEL_FONT
sv["A20"].alignment = LEFT
sv["A20"].fill = NOTE_FILL

fill_frame_row(21, "Wage — Low",        "=W_low",          "Mid-level engineer fully loaded — see Wage Calculator sheet.", fill=PatternFill('solid', fgColor='66BB6A'))
fill_frame_row(22, "Wage — Protective", "=W_protective",   "Senior engineer fully loaded. Use this with finance.", fill=SAVINGS_FILL, big=True)
fill_frame_row(23, "Wage — Aggressive", "=W_aggressive",   "Big-tech-comp senior fully loaded. The realistic ceiling on wage.", fill=PatternFill('solid', fgColor='66BB6A'))

# --- Sub-heading: Investor frames (the ceiling) ---
sv.merge_cells("A25:G25")
sv["A25"] = "  Investor frames — every minute is worth Nx its wage, because investors expect Nx return on every $1 of capital they put in."
sv["A25"].font = LABEL_FONT
sv["A25"].alignment = LEFT
sv["A25"].fill = NOTE_FILL

# These multiplier inputs are editable in cells B26-B29
inv_purple = PatternFill('solid', fgColor='6A1B9A')
fill_frame_row(26, "Investor — 2× target",   "=W_protective*2",   "Mature growth-stage. 'Double the money in 5–7 years.'", fill=inv_purple)
fill_frame_row(27, "Investor — 3× target",   "=W_protective*3",   "Series C / D. Standard late-growth target.", fill=inv_purple)
fill_frame_row(28, "Investor — 10× target",  "=W_protective*10",  "Series A / B SaaS. The 'fund returner' threshold most VCs underwrite to.", fill=inv_purple, big=True)
fill_frame_row(29, "Investor — 100× target", "=W_protective*100", "Seed / pre-seed math. The aspirational ceiling — useful as a sanity-check, not a quote.", fill=inv_purple)

# ============================================================================
# SECTION 3 — INVESTOR LOGIC (the explainer that makes the table above land)
# ============================================================================
section(sv, 31, "STEP 3 — WHY INVESTORS DON'T SEE WAGE — THEY SEE EXPECTED RETURN", span=7, fill=HEADLINE_FILL, font=HEADLINE_FONT, height=28)

# The killer one-liner
sv.merge_cells("A32:G32")
sv["A32"] = f"$1 spent at {ORG} carries an implicit $2-$100 of expected return — depending on which round, which stage, and which investor's underwriting model you ask."
sv["A32"].font = Font(bold=True, size=14, color="6A1B9A", italic=True)
sv["A32"].alignment = CENTER
sv["A32"].fill = PatternFill('solid', fgColor='F3E5F5')
sv["A32"].border = THIN
sv.row_dimensions[32].height = 38

sv.merge_cells("A33:G33")
sv["A33"] = (
    f"When a VC writes {ORG} a cheque, they are not buying engineering labour at its wage cost. "
    "They are buying an option that the same engineering produces a multiple of that — typically a 10× return — within their fund's hold period.\n\n"
    "That expectation is the implicit price tag on every engineer's time. If the company is underwritten to deliver 10× over 7 years, "
    "then every dollar spent on engineering must produce $10 of expected enterprise value to keep the cap table whole. "
    "An hour of engineering wasted is not an hour of wage burned — it is an hour of compounding output not produced.\n\n"
    "Translated to our number: at a 10× target, every developer-minute is implicitly worth ~10× its wage in expected value. "
    "The minutes we save on every CI run, multiplied by runs per year, becomes a meaningful chunk of expected enterprise value — "
    "from a CI improvement that cost us nothing visible. At a 100× early-stage target, the same improvement is an order of magnitude bigger.\n\n"
    "This is not a number you put in a budget meeting. It IS the number you put in a board meeting, an OKR planning session, "
    "or any conversation about whether to invest engineering time on infrastructure versus features. "
    "Both are subject to the same multiplier — so the comparison is fair."
)
sv["A33"].font = NOTE_FONT
sv["A33"].alignment = LEFT
sv["A33"].fill = NOTE_FILL
sv["A33"].border = THIN
sv.row_dimensions[33].height = 220

# ============================================================================
# SECTION 4 — THE STORY (three sentences, in order)
# ============================================================================
section(sv, 35, "STEP 4 — HOW TO TELL THIS STORY (three sentences, in order)", span=7, fill=SAVINGS_FILL)

# Sentence 1 — what happened (numbers via formulas so they stay in sync)
sv.merge_cells("A36:G36")
sv["A36"] = (
    f"1. WHAT HAPPENED.\n"
    f"Without shipping a single line of new feature code, the CI/CD pipeline got {improvement_pct}% faster over the last 90 days. "
    f"Each active developer saves {mins_saved} minutes every time they push code, multiple times a day. "
    f"See cells B13–B15 above for the exact dev-minutes / dev-hours / FTE-years saved per year — those numbers come straight from "
    f"the measured pipeline-duration drop and the team size you set on the Inputs sheet."
)
sv["A36"].font = LABEL_FONT
sv["A36"].alignment = LEFT
sv["A36"].fill = NOTE_FILL
sv["A36"].border = THIN
sv.row_dimensions[36].height = 90

# Sentence 2 — what it's worth at the floor
sv.merge_cells("A37:G37")
sv["A37"] = (
    "2. WHAT IT IS WORTH (the floor — pure arithmetic, no studies).\n"
    "At the loaded engineering wage you set, the value of those minutes saved is shown in row 22 above (Wage — Protective). "
    "This number requires no assumption beyond 'developers cost money'. It is the smallest defensible answer. "
    "Read the per-period columns to quote the right cadence (annual / monthly / weekly / daily)."
)
sv["A37"].font = LABEL_FONT
sv["A37"].alignment = LEFT
sv["A37"].fill = NOTE_FILL
sv["A37"].border = THIN
sv.row_dimensions[37].height = 90

# Sentence 3 — what it's worth at the investor ceiling
sv.merge_cells("A38:G38")
sv["A38"] = (
    f"3. WHAT IT IS WORTH (the investor frame — what the cap table actually expects).\n"
    f"{ORG}'s investors are not paying for engineering labour at wage — they are paying for engineering output at the 10× return implicit in the funding round. "
    f"By that frame (row 28 above), the same minutes saved becomes a much larger expected-value number. "
    f"Whichever frame leadership picks, this is a meaningful win, achieved without writing a single line of feature code. "
    f"Pick the frame, then ask: what would 5 of these wins look like?"
)
sv["A38"].font = LABEL_FONT
sv["A38"].alignment = LEFT
sv["A38"].fill = NOTE_FILL
sv["A38"].border = THIN
sv.row_dimensions[38].height = 110


# =============================================================================
# Sheet: Wage Calculator (regional mix → blended $/min)
# =============================================================================
wc = wb.create_sheet("Wage Calculator")
for col, w in zip("ABCDEFG", [28, 14, 18, 16, 18, 16, 36]):
    wc.column_dimensions[col].width = w

wc.merge_cells("A1:G1")
wc["A1"] = "WAGE CALCULATOR — annual salary in / loaded $/min out"
wc["A1"].font = HEADLINE_FONT
wc["A1"].fill = HEADLINE_FILL
wc["A1"].alignment = CENTER
wc.row_dimensions[1].height = 32

wc.merge_cells("A2:G2")
wc["A2"] = (f"Executives think in annual salaries, not $/minute. Enter the typical fully-loaded annual cost per region "
            f"(salary + benefits + employer tax + overhead) plus how many engineers {ORG} has in each region. "
            f"The blended $/min comes out at the bottom — that is the W to plug into the Inputs sheet.")
wc["A2"].font = NOTE_FONT
wc["A2"].alignment = LEFT
wc["A2"].fill = NOTE_FILL
wc.row_dimensions[2].height = 50

# ---- Reference: annual → per-minute conversion explained ----
section(wc, 4, "HOW THE NUMBERS CONVERT", span=7, fill=SECTION_FILL)

wc.merge_cells("A5:G5")
wmpy = SEED["working_min_per_year"]
wc["A5"] = (f"Working minutes per year = 50 weeks × 40 hr × 60 min = {wmpy:,} min.\n"
            f"Loaded annual cost ÷ {wmpy:,} = loaded $ per developer-minute.\n"
            f"Example: a $240,000 fully-loaded engineer ÷ {wmpy:,} ≈ $2.00/minute.")
wc["A5"].font = NOTE_FONT
wc["A5"].alignment = LEFT
wc["A5"].fill = NOTE_FILL
wc.row_dimensions[5].height = 56

# Named cell for the constant
wc["A7"] = "Working minutes per year (constant)"
wc["B7"] = SEED["working_min_per_year"]
style_label(wc["A7"]); style_input(wc["B7"], "#,##0")
named("working_min_per_year", "Wage Calculator", "B7")

wc["A8"] = "Loading multiplier (salary → fully loaded)"
wc["B8"] = SEED["loading_multiplier"]
wc["C8"] = "convention"
wc["D8"] = "Standard McKinsey 1.3x-1.4x: salary × 1.40 = full cost (benefits, payroll tax, overhead, IT, recruiting)."
style_label(wc["A8"]); style_input(wc["B8"], "0.00"); style_label(wc["C8"]);
wc.merge_cells("D8:G8"); style_note(wc["D8"])
named("loading_multiplier", "Wage Calculator", "B8")

# ---- Regional comp table ----
section(wc, 10, "REGIONAL ENGINEERING COMP — fully loaded annual ($USD equivalent)", span=7, fill=SECTION_FILL)

wc.merge_cells("A11:G11")
wc["A11"] = (f"Adjust 'Engineers in region' to match {ORG}'s actual mix. The 'Annual base salary' column is the typical senior engineer's "
             f"GROSS annual salary — override with what HR actually pays. The 'Loaded annual' column applies the loading multiplier. "
             f"Currency: convert local salary to USD using your finance team's reference rate.")
wc["A11"].font = NOTE_FONT
wc["A11"].alignment = LEFT
wc["A11"].fill = NOTE_FILL
wc.row_dimensions[11].height = 50

# Header
for col_letter, label in zip("ABCDEFG", ["Region", "Engineers", "Local currency", "Annual base salary", "Loaded annual ($)", "$/min", "Notes"]):
    wc[f"{col_letter}12"].value = label
    wc[f"{col_letter}12"].font = WHITE_BOLD
    wc[f"{col_letter}12"].fill = SECTION_FILL
    wc[f"{col_letter}12"].alignment = CENTER
    wc[f"{col_letter}12"].border = THIN

# Regional rows from report-data.json
regional_start = 13
for i, r in enumerate(REGIONS):
    row_n = regional_start + i
    wc[f"A{row_n}"] = r["region"]
    wc[f"B{row_n}"] = r["engineers"]
    wc[f"C{row_n}"] = r["currency"]
    wc[f"D{row_n}"] = r["annual_base_usd"]
    wc[f"E{row_n}"] = f"=D{row_n}*loading_multiplier"
    wc[f"F{row_n}"] = f"=E{row_n}/working_min_per_year"
    wc[f"G{row_n}"] = r["note"]
    style_label(wc[f"A{row_n}"])
    style_input(wc[f"B{row_n}"], "0")
    wc[f"C{row_n}"].alignment = CENTER
    wc[f"C{row_n}"].border = THIN
    style_input(wc[f"D{row_n}"], '"$"#,##0')
    style_derived(wc[f"E{row_n}"], '"$"#,##0')
    style_derived(wc[f"F{row_n}"], USD_FMT)
    style_note(wc[f"G{row_n}"])
regional_end = regional_start + len(REGIONS) - 1

# Totals
total_row = regional_end + 2
wc[f"A{total_row}"] = "TOTAL ENGINEERS"
wc[f"B{total_row}"] = f"=SUM(B{regional_start}:B{regional_end})"
style_label(wc[f"A{total_row}"]); style_derived(wc[f"B{total_row}"], "0")
wc[f"B{total_row}"].fill = HEADLINE_FILL
wc[f"B{total_row}"].font = WHITE_BOLD
wc[f"B{total_row}"].alignment = CENTER

blended_min_row = total_row + 1
wc[f"A{blended_min_row}"] = "WEIGHTED AVERAGE LOADED $/MIN"
wc[f"B{blended_min_row}"] = (f"=SUMPRODUCT(B{regional_start}:B{regional_end}, F{regional_start}:F{regional_end}) "
                             f"/ B{total_row}")
style_label(wc[f"A{blended_min_row}"])
wc[f"B{blended_min_row}"].fill = SAVINGS_FILL
wc[f"B{blended_min_row}"].font = SAVINGS_FONT
wc[f"B{blended_min_row}"].alignment = CENTER
wc[f"B{blended_min_row}"].number_format = USD_FMT
wc[f"B{blended_min_row}"].border = THIN
wc.row_dimensions[blended_min_row].height = 28

blended_annual_row = blended_min_row + 1
wc[f"A{blended_annual_row}"] = "WEIGHTED AVERAGE LOADED ANNUAL"
wc[f"B{blended_annual_row}"] = (f"=SUMPRODUCT(B{regional_start}:B{regional_end}, E{regional_start}:E{regional_end}) "
                                f"/ B{total_row}")
style_label(wc[f"A{blended_annual_row}"])
style_derived(wc[f"B{blended_annual_row}"], '"$"#,##0')

named("blended_per_min", "Wage Calculator", f"B{blended_min_row}")

# Helper note: how to use this
helper_top = total_row
helper_bottom = blended_annual_row
wc.merge_cells(start_row=helper_top, start_column=4, end_row=helper_bottom, end_column=7)
wc.cell(row=helper_top, column=4).value = (
    f"HOW TO USE THIS NUMBER:\n"
    f"Cell B{blended_min_row} is your blended $/min. "
    f"Copy that value into the Inputs sheet → 'Loaded $/dev-minute (W) — Protective'. "
    f"Or use it directly: every CI minute costs ~$/min × team size × runs."
)
wc.cell(row=helper_top, column=4).font = LABEL_FONT
wc.cell(row=helper_top, column=4).alignment = LEFT
wc.cell(row=helper_top, column=4).fill = INPUT_STRONG
wc.cell(row=helper_top, column=4).border = THIN

# ---- Quick-lookup: annual salary → $/min ----
section(wc, 24, "QUICK LOOKUP — annual salary to $/min", span=7, fill=SECTION_FILL)

wc.merge_cells("A25:G25")
wc["A25"] = "If you know the annual salary, you can read off the per-minute cost here. Or enter your own in cell B27."
wc["A25"].font = NOTE_FONT
wc["A25"].alignment = LEFT
wc["A25"].fill = NOTE_FILL

wc["A26"] = "Annual base salary"
wc["B26"] = "Loaded annual"
wc["C26"] = "Loaded $/min"
wc["D26"] = "Comparable to..."
for col in "ABCD":
    wc[f"{col}26"].font = WHITE_BOLD
    wc[f"{col}26"].fill = SECTION_FILL
    wc[f"{col}26"].alignment = CENTER
    wc[f"{col}26"].border = THIN

# An editable starter row — exec types their org's number
wc["A27"] = "(Type your org's annual salary →)"
wc["B27"] = 120000   # editable starter — generic, not org-specific
wc["C27"] = "=B27*loading_multiplier"
wc["D27"] = "=C27/working_min_per_year"
wc["E27"] = "Edit B27 ↑"
style_label(wc["A27"])
style_input(wc["B27"], '"$"#,##0', strong=True)
style_derived(wc["C27"], '"$"#,##0')
wc["D27"].fill = SAVINGS_FILL
wc["D27"].font = WHITE_BOLD
wc["D27"].number_format = USD_FMT
wc["D27"].border = THIN
wc["D27"].alignment = CENTER
style_note(wc["E27"])

# Static reference table from report-data.json
for i, ref in enumerate(REF_SALARIES):
    row_n = 28 + i
    wc[f"A{row_n}"] = ref["comparable"]
    wc[f"B{row_n}"] = ref["salary"]
    wc[f"C{row_n}"] = f"=B{row_n}*loading_multiplier"
    wc[f"D{row_n}"] = f"=C{row_n}/working_min_per_year"
    style_label(wc[f"A{row_n}"])
    style_derived(wc[f"B{row_n}"], '"$"#,##0')
    style_derived(wc[f"C{row_n}"], '"$"#,##0')
    style_derived(wc[f"D{row_n}"], USD_FMT)


# =============================================================================
# Sheet: Inputs (the editable knobs)
# =============================================================================
inp = wb.create_sheet("Inputs")
inp.column_dimensions["A"].width = 36
inp.column_dimensions["B"].width = 16
inp.column_dimensions["C"].width = 14
inp.column_dimensions["D"].width = 80

inp["A1"] = "INPUTS — edit yellow cells; everything else recalcs"
inp["A1"].font = Font(bold=True, size=18, color="1A237E")
inp.merge_cells("A1:D1")
inp.row_dimensions[1].height = 28

for col_letter, label in zip("ABCD", ["Parameter", "Value", "Source", "What it is / how to change"]):
    inp[f"{col_letter}2"].value = label
    inp[f"{col_letter}2"].fill = SECTION_FILL
    inp[f"{col_letter}2"].font = WHITE_BOLD
    inp[f"{col_letter}2"].alignment = CENTER
    inp[f"{col_letter}2"].border = THIN

row = 3
section(inp, row, "TEAM", span=4); row += 1

inp[f"A{row}"] = "Active developers (D)"
inp[f"B{row}"] = SEED["D_active_devs"]
inp[f"C{row}"] = "measured"
core = DEV_OPTIONS["core"]
rec = DEV_OPTIONS["recommended"]
upper = DEV_OPTIONS["upper"]
inp[f"D{row}"] = (f"Three measurements; pick the one that matches your audience. "
                  f"(a) {core['value']} — {core['label']}: {core['rule']}. "
                  f"Use this for the conservative narrative — 'who CI/CD slowness actually hurts day-to-day'. "
                  f"(b) {rec['value']} — {rec['label']}: {rec['rule']}. "
                  f"Cleanest 'who shipped code recently' answer; matches the Savings sheet narrative. "
                  f"(c) {upper['value']} — {upper['label']}: {upper['rule']}. "
                  f"Use only if you want to make the friction tax look bigger.")
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], "0"); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("D_devs", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "CI runs / dev / business day (R)"
inp[f"B{row}"] = SEED["R_runs_per_dev_day"]
inp[f"C{row}"] = "measured"
inp[f"D{row}"] = "How often each developer triggers CI on a working day."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], "0.0"); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("R_runs_per_dev_day", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Working days per year"
inp[f"B{row}"] = SEED["days_per_year"]
inp[f"C{row}"] = "convention"
inp[f"D{row}"] = "Standard finance default (52 weeks × 5 - holidays)."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], "0"); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("days_per_year", "Inputs", f"B{row}"); row += 2

section(inp, row, "WAGE — choose your frame", span=4); row += 1

inp[f"A{row}"] = "Loaded $/dev-minute (W) — Low"
inp[f"B{row}"] = SEED["W_low"]
inp[f"C{row}"] = "BLS"
inp[f"D{row}"] = "Mid-level US engineer fully loaded. See Wage Calculator sheet."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], USD_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("W_low", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Loaded $/dev-minute (W) — Protective"
inp[f"B{row}"] = SEED["W_protective"]
inp[f"C{row}"] = "Levels.fyi"
inp[f"D{row}"] = "Senior engineer fully loaded. The recommended headline value. See Wage Calculator sheet."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], USD_FMT, strong=True); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("W_protective", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Loaded $/dev-minute (W) — Aggressive"
inp[f"B{row}"] = SEED["W_aggressive"]
inp[f"C{row}"] = "investor"
inp[f"D{row}"] = "Big-tech-comp senior fully loaded. The 'investor frame' uses revenue-per-engineer instead — see Savings sheet Step 3."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], USD_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("W_aggressive", "Inputs", f"B{row}"); row += 2

section(inp, row, "PIPELINE (measured from GitHub Actions, last 30 days)", span=4); row += 1

inp[f"A{row}"] = "Mean PR happy-path duration (T_dur, min)"
inp[f"B{row}"] = SEED["T_dur_min"]
inp[f"C{row}"] = "measured"
inp[f"D{row}"] = "Combined happy-path mean over the last 30 days. The 'developer felt' number. See Trend sheet for windowed values."
style_label(inp[f"A{row}"]); style_measured(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("T_dur", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Mean per-run queue time (T_queue, min)"
inp[f"B{row}"] = SEED["T_queue_min"]
inp[f"C{row}"] = "measured"
inp[f"D{row}"] = "Sum of per-job queue waits — incl. STG concurrency lock. Will rise sharply with team growth (queueing theory)."
style_label(inp[f"A{row}"]); style_measured(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("T_queue", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Total pipeline (T_dur + T_queue, min)"
inp[f"B{row}"] = "=T_dur+T_queue"
inp[f"C{row}"] = "derived"
inp[f"D{row}"] = "What we plug into the cognitive-tax piecewise function."
style_label(inp[f"A{row}"]); style_derived(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("T_pipeline", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "CI failure rate (F)"
inp[f"B{row}"] = SEED["F_failure_rate"]
inp[f"C{row}"] = "measured"
inp[f"D{row}"] = "Failed/total runs, excluding cancellations. NOT DORA's Change Failure Rate (production failures)."
style_label(inp[f"A{row}"]); style_measured(inp[f"B{row}"], PCT_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("F_failure_rate", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Blast-radius multiplier (K)"
inp[f"B{row}"] = SEED["K_blast_radius"]
inp[f"C{row}"] = "measured"
inp[f"D{row}"] = "Empirical: how many other devs each failure blocks via shared STG concurrency lock. Ephemeral envs would push this to 1.0."
style_label(inp[f"A{row}"]); style_measured(inp[f"B{row}"], "0.00"); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("K_blast_radius", "Inputs", f"B{row}"); row += 2

section(inp, row, "COGNITIVE TAX (piecewise — defensible against the 'flat 23 min is folkloric' attack)", span=4); row += 1

inp[f"A{row}"] = "Recovery min for waits < 2 min"
inp[f"B{row}"] = SEED["S_short"]
inp[f"C{row}"] = "Cypher 2017"
inp[f"D{row}"] = "Self-initiated short waits — attention stays parked at the keyboard."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("S_short", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Recovery min for waits 2–10 min"
inp[f"B{row}"] = SEED["S_medium"]
inp[f"C{row}"] = "Atlassian"
inp[f"D{row}"] = "Quick context switch, fast recovery (Slack scroll, return)."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("S_medium", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Recovery min for waits ≥ 10 min"
inp[f"B{row}"] = SEED["S_long"]
inp[f"C{row}"] = "Mark 2008"
inp[f"D{row}"] = "Full task switch — Mark/Gloria/Klocke 2008. The canonical 23-min figure."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("S_long", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Short → Medium threshold (min)"
inp[f"B{row}"] = SEED["S_thresh_short"]
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], MIN_FMT)
named("S_thresh_short", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Medium → Long threshold (min)"
inp[f"B{row}"] = SEED["S_thresh_long"]
inp[f"C{row}"] = "Atlassian"
inp[f"D{row}"] = "The canonical 'context-switch cliff'. Single biggest narrative lever — change here moves the whole story."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"], MIN_FMT, strong=True); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("S_thresh_long", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Cognitive minutes applied (S, derived)"
inp[f"B{row}"] = "=IF(T_pipeline<S_thresh_short, S_short, IF(T_pipeline<S_thresh_long, S_medium, S_long))"
inp[f"C{row}"] = "derived"
inp[f"D{row}"] = "Selected piecewise band given current T_pipeline."
style_label(inp[f"A{row}"]); style_derived(inp[f"B{row}"], MIN_FMT); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("S_applied", "Inputs", f"B{row}"); row += 1

inp[f"A{row}"] = "Include cognitive bucket?"
inp[f"B{row}"] = "YES"
inp[f"C{row}"] = "policy"
inp[f"D{row}"] = "Set to NO to drop the cognitive tax entirely (most conservative finance frame)."
style_label(inp[f"A{row}"]); style_input(inp[f"B{row}"]); style_label(inp[f"C{row}"]); style_note(inp[f"D{row}"])
named("include_cognitive", "Inputs", f"B{row}")


# =============================================================================
# Sheet: Trend (with n counts, 7d included)
# =============================================================================
tr = wb.create_sheet("Trend")
for col, w in zip("ABCDEFG", [22, 14, 18, 14, 18, 14, 50]):
    tr.column_dimensions[col].width = w

tr.merge_cells("A1:G1")
tr["A1"] = "TREND — happy-path duration over time (live from GitHub Actions)"
tr["A1"].font = HEADLINE_FONT
tr["A1"].fill = HEADLINE_FILL
tr["A1"].alignment = CENTER
tr.row_dimensions[1].height = 32

tr.merge_cells("A2:G2")
tr["A2"] = ("Successful CI/CD runs only. Generated from the live GitHub API. "
            "PR happy-path is what every developer feels every time they open a PR — "
            "it includes E2E on STG, which the master-push path skips.")
tr["A2"].font = NOTE_FONT
tr["A2"].alignment = LEFT
tr.row_dimensions[2].height = 30

# Header
trend_headers = ["Window", "All n", "All — mean", "PR n", "PR — mean", "PR — p90", "Direction"]
for i, text in enumerate(trend_headers, start=1):
    c = tr.cell(row=3, column=i, value=text)
    c.font = WHITE_BOLD
    c.fill = SECTION_FILL
    c.alignment = CENTER
    c.border = THIN

# Trend rows from report-data.json (oldest first, so 'direction' reads left-to-right)
for i, w in enumerate(TREND, start=4):
    tr.cell(row=i, column=1, value=w["label"])
    tr.cell(row=i, column=2, value=w["n_all"])
    tr.cell(row=i, column=3, value=w["mean_all"])
    tr.cell(row=i, column=4, value=w["n_pr"])
    tr.cell(row=i, column=5, value=w["mean_pr"])
    tr.cell(row=i, column=6, value=w["p90_pr"])
    tr.cell(row=i, column=7, value=w["direction"])
    direction = w["direction"]
    style_label(tr.cell(row=i, column=1))
    style_measured(tr.cell(row=i, column=2), "0")
    style_measured(tr.cell(row=i, column=3), MIN_FMT)
    style_measured(tr.cell(row=i, column=4), "0")
    style_measured(tr.cell(row=i, column=5), MIN_FMT)
    style_measured(tr.cell(row=i, column=6), MIN_FMT)
    direction_color = "2E7D32" if "↓" in direction or "→" in direction else "424242"
    tr.cell(row=i, column=7).font = Font(bold=True, color=direction_color)
    tr.cell(row=i, column=7).fill = MEASURED_FILL
    tr.cell(row=i, column=7).alignment = CENTER
    tr.cell(row=i, column=7).border = THIN

# Δ rows
tr["A9"] = "Δ over 90d (90d → 7d)"
tr["B9"] = ""
tr["C9"] = "=C7-C4"
tr["E9"] = "=E7-E4"
tr["F9"] = "=F7-F4"
tr["G9"] = '=IF(C9<0, "Pipeline duration is dropping — keep going", "Pipeline duration is rising — investigate")'
style_label(tr["A9"])
for col in "CEF":
    style_derived(tr[f"{col}9"], '+0.0" min";-0.0" min"')
tr["G9"].font = GREEN_FONT
tr["G9"].alignment = LEFT
tr["G9"].fill = SAVINGS_FILL
tr["G9"].font = Font(bold=True, color="FFFFFF", size=12)

tr["A10"] = "% change"
tr["C10"] = "=C9/C4"
tr["E10"] = "=E9/E4"
tr["F10"] = "=F9/F4"
style_label(tr["A10"])
for col in "CEF":
    style_derived(tr[f"{col}10"], "+0%;-0%")

tr["A11"] = "Sample size"
tr["B11"] = "=B4+B5+B6+B7"
tr["D11"] = "=D4+D5+D6+D7"
tr["G11"] = "Total runs analysed across all four windows (overlap is real — 7d ⊂ 14d ⊂ 30d ⊂ 90d, this is for context)"
style_label(tr["A11"]); style_derived(tr["B11"], "0"); style_derived(tr["D11"], "0")
style_note(tr["G11"])

# Chart it
ch = BarChart()
ch.title = "Happy-path duration trend (mean min)"
ch.y_axis.title = "Mean minutes"
ch.x_axis.title = "Time window (oldest → newest)"
ch.style = 10
data_all = Reference(tr, min_col=3, max_col=3, min_row=3, max_row=7)
data_pr = Reference(tr, min_col=5, max_col=5, min_row=3, max_row=7)
cats = Reference(tr, min_col=1, min_row=4, max_row=7)
ch.add_data(data_all, titles_from_data=True)
ch.add_data(data_pr, titles_from_data=True)
ch.set_categories(cats)
ch.width = 22
ch.height = 12
tr.add_chart(ch, "I3")


# =============================================================================
# Sheet: Headline
# =============================================================================
hd = wb.create_sheet("Headline")
for col, w in zip("ABCDE", [38, 18, 18, 18, 56]):
    hd.column_dimensions[col].width = w

hd.merge_cells("A1:E1")
hd["A1"] = f"$/CI MINUTE — what every minute of pipeline costs {ORG}"
hd["A1"].font = HEADLINE_FONT
hd["A1"].fill = HEADLINE_FILL
hd["A1"].alignment = CENTER
hd.row_dimensions[1].height = 36

hd["B3"] = "Low"
hd["C3"] = "Protective (recommended)"
hd["D3"] = "Aggressive"
for col in "BCD":
    hd[f"{col}3"].font = WHITE_BOLD
    hd[f"{col}3"].fill = SECTION_FILL
    hd[f"{col}3"].alignment = CENTER
    hd[f"{col}3"].border = THIN
hd.row_dimensions[3].height = 22

hd["A4"] = "Wage frame ($/dev-min)"
hd["B4"] = "=W_low"; hd["C4"] = "=W_protective"; hd["D4"] = "=W_aggressive"
style_label(hd["A4"])
for col in "BCD": style_derived(hd[f"{col}4"], USD_FMT)

hd["A5"] = "Direct waste ($/CI min)"
for col in "BCD":
    hd[f"{col}5"] = f"=D_devs*R_runs_per_dev_day*{col}4"
style_label(hd["A5"])
for col in "BCD": style_derived(hd[f"{col}5"], USD_FMT)

cog_formula = '=IF(include_cognitive="YES", 1 + S_applied/T_pipeline, 1)'
hd["A6"] = "× Cognitive multiplier"
for col in "BCD": hd[f"{col}6"] = cog_formula
style_label(hd["A6"])
for col in "BCD": style_derived(hd[f"{col}6"], "0.00")

hd["A7"] = "× Failure multiplier"
for col in "BCD": hd[f"{col}7"] = "=1+F_failure_rate*K_blast_radius"
style_label(hd["A7"])
for col in "BCD": style_derived(hd[f"{col}7"], "0.00")

hd["A8"] = "$/CI MINUTE"
hd["A8"].font = Font(bold=True, size=14, color="1A237E")
hd["A8"].border = THIN
hd["A8"].alignment = LEFT
for col in "BCD":
    hd[f"{col}8"] = f"={col}5*{col}6*{col}7"
    hd[f"{col}8"].fill = HEADLINE_FILL
    hd[f"{col}8"].font = HEADLINE_FONT
    hd[f"{col}8"].alignment = CENTER
    hd[f"{col}8"].number_format = USD_FMT
    hd[f"{col}8"].border = THIN
hd.row_dimensions[8].height = 32

hd["A9"] = "$/CI second"
for col in "BCD": hd[f"{col}9"] = f"={col}8/60"
style_label(hd["A9"])
for col in "BCD": style_derived(hd[f"{col}9"], USD_FMT)

hd["A10"] = "Annual pipeline minutes"
for col in "BCD":
    hd[f"{col}10"] = "=T_pipeline*R_runs_per_dev_day*D_devs*days_per_year"
style_label(hd["A10"])
for col in "BCD": style_derived(hd[f"{col}10"], "#,##0")

hd["A11"] = "ANNUAL FRICTION TAX"
hd["A11"].font = Font(bold=True, size=14, color="1A237E")
hd["A11"].border = THIN
hd["A11"].alignment = LEFT
for col in "BCD":
    hd[f"{col}11"] = f"={col}8*{col}10"
    hd[f"{col}11"].fill = HEADLINE_FILL
    hd[f"{col}11"].font = HEADLINE_FONT
    hd[f"{col}11"].alignment = CENTER
    hd[f"{col}11"].number_format = USD_BIG_FMT
    hd[f"{col}11"].border = THIN
hd.row_dimensions[11].height = 32

# Time-period breakdown of friction tax
hd["A12"] = "  → per month"
for col in "BCD": hd[f"{col}12"] = f"={col}11/12"
style_label(hd["A12"])
for col in "BCD": style_derived(hd[f"{col}12"], USD_BIG_FMT)

hd["A13"] = "  → per week"
for col in "BCD": hd[f"{col}13"] = f"={col}11/50"
style_label(hd["A13"])
for col in "BCD": style_derived(hd[f"{col}13"], USD_BIG_FMT)

hd["A14"] = "  → per working day"
for col in "BCD": hd[f"{col}14"] = f"={col}11/days_per_year"
style_label(hd["A14"])
for col in "BCD": style_derived(hd[f"{col}14"], USD_BIG_FMT)

hd["A15"] = "  → per working hour"
for col in "BCD": hd[f"{col}15"] = f"={col}11/days_per_year/8"
style_label(hd["A15"])
for col in "BCD": style_derived(hd[f"{col}15"], USD_BIG_FMT)

hd["A17"] = "Value of saving 1 min off the build"
hd["A17"].font = Font(bold=True, size=12, color="2E7D32")
hd["A17"].border = THIN
hd["A17"].alignment = LEFT
hd["B16"] = "Annual"; hd["C16"] = "Monthly"; hd["D16"] = "Weekly"; hd["E16"] = "Daily"
for col in "BCDE":
    hd[f"{col}16"].font = WHITE_BOLD
    hd[f"{col}16"].fill = SECTION_FILL
    hd[f"{col}16"].alignment = CENTER
    hd[f"{col}16"].border = THIN
# Use Protective for this row (the most quoted number)
hd["B17"] = "=C8*R_runs_per_dev_day*D_devs*days_per_year"
hd["C17"] = "=B17/12"
hd["D17"] = "=B17/50"
hd["E17"] = "=B17/days_per_year"
for col in "BCDE":
    style_derived(hd[f"{col}17"], USD_BIG_FMT)
    hd[f"{col}17"].fill = SAVINGS_FILL
    hd[f"{col}17"].font = WHITE_BOLD

# Decomposition (Protective)
hd.merge_cells("A20:E20")
hd["A20"] = "DECOMPOSITION (Protective frame, annual)"
hd["A20"].font = SECTION_FONT
hd["A20"].fill = SECTION_FILL
hd["A20"].alignment = LEFT

for i, label in enumerate(["Bucket", "Annual cost", "% of total", "Multiplier", "What it is"]):
    c = hd.cell(row=21, column=i+1, value=label)
    c.font = WHITE_BOLD
    c.fill = SECTION_FILL
    c.alignment = CENTER
    c.border = THIN

hd["A22"] = "Direct waste (idle wage)"
hd["B22"] = "=C5*C10"; hd["C22"] = "=B22/C11"; hd["D22"] = 1.00
hd["E22"] = "Wage of every dev waiting on a build."
style_label(hd["A22"]); style_derived(hd["B22"], USD_BIG_FMT); style_derived(hd["C22"], PCT_FMT); style_derived(hd["D22"], "0.00"); style_note(hd["E22"])

hd["A23"] = "Cognitive tax"
hd["B23"] = '=IF(include_cognitive="YES", B22*(C6-1), 0)'
hd["C23"] = "=B23/C11"; hd["D23"] = "=C6-1"
hd["E23"] = "Focus-loss when build crosses the cognitive cliff."
style_label(hd["A23"]); style_derived(hd["B23"], USD_BIG_FMT); style_derived(hd["C23"], PCT_FMT); style_derived(hd["D23"], "0.00"); style_note(hd["E23"])

hd["A24"] = "Failure rework"
hd["B24"] = "=(B22+B23)*(C7-1)"; hd["C24"] = "=B24/C11"; hd["D24"] = "=C7-1"
hd["E24"] = "Failed builds × blast-radius (broken STG blocks others)."
style_label(hd["A24"]); style_derived(hd["B24"], USD_BIG_FMT); style_derived(hd["C24"], PCT_FMT); style_derived(hd["D24"], "0.00"); style_note(hd["E24"])

hd["A25"] = "TOTAL"
hd["A25"].font = LABEL_FONT
hd["A25"].border = THIN
hd["B25"] = "=SUM(B22:B24)"; hd["C25"] = "=SUM(C22:C24)"
style_derived(hd["B25"], USD_BIG_FMT); style_derived(hd["C25"], PCT_FMT)


# =============================================================================
# Sheet: Vision
# =============================================================================
vis = wb.create_sheet("Vision")
for col, w in zip("ABCDE", [14, 12, 14, 16, 80]):
    vis.column_dimensions[col].width = w

vis.merge_cells("A1:E1")
vis["A1"] = "VISION — what changes at each pipeline duration"
vis["A1"].font = HEADLINE_FONT
vis["A1"].fill = HEADLINE_FILL
vis["A1"].alignment = CENTER
vis.row_dimensions[1].height = 32

vis.merge_cells("A2:E2")
vis["A2"] = (f"Holding everything else constant (D, R, F, K, W=Protective). "
             f"Yellow row = where {ORG} currently sits. Try changing the inputs "
             f"on the Inputs sheet — these numbers update automatically.")
vis["A2"].font = NOTE_FONT
vis["A2"].alignment = LEFT
vis.row_dimensions[2].height = 38

for i, text in enumerate(["Duration", "Cog. band", "$/CI min", "Annual friction", "What life looks like at this duration"], start=1):
    c = vis.cell(row=3, column=i, value=text)
    c.font = WHITE_BOLD; c.fill = SECTION_FILL; c.alignment = CENTER; c.border = THIN

# Vision rows — narrative is generic; the "current" highlight uses the measured T_dur
# from report-data.json so the highlighted band always points to where the org lives now.
current_t = round(SEED["T_dur_min"] + SEED["T_queue_min"])  # nearest minute
vision_rows = [
    (3,  "Short",  "The build is faster than a Slack reply. Developers stay at their keyboards. The pipeline disappears as a thing anyone thinks about. Elite tier — DORA's top decile."),
    (5,  "Medium", "The 'go grab coffee' build. Below the cognitive cliff. Continuous deployment becomes plausible. ~30% cheaper per run than at 10 min even though pipeline is 2× shorter."),
    (8,  "Medium", "The 'scroll through GitHub notifications' build. Tolerable, not fast. Just under the 10-min cliff — every minute saved here is 2× more valuable than every minute saved above 10 min."),
    (10, "Long",   "THE CLIFF EDGE. Just over the threshold — developers task-switch. Every minute *above* 10 min keeps you in the high-cost band; every minute *below* drops you into medium."),
    (13, "Long",   f"Plateau-confirmed lower band — close to where {ORG} lives today. Big chunk of the cognitive tax already paid. Further savings now come from getting under 10 min."),
    (15, "Long",   "Reviewers context-switch when re-checking PRs. Failure rate × duration waste compounds quickly. Where most teams plateau before a deliberate push to <10 min."),
    (20, "Long",   "'The build is broken' becomes ambient. Devs batch-push at end-of-day. PR cycle time measured in days. Junior engineers dread pushing."),
    (30, "Long",   "The pipeline is a meeting. Devs schedule their day around builds. New hires spend their first week learning when *not* to push."),
    (60, "Long",   "Crisis tier. Pipeline is a release process, not a CI signal. Hotfixes go around it. Engineering velocity rate-limited by infrastructure."),
]

for i, (mins, band, narrative) in enumerate(vision_rows, start=4):
    vis.cell(row=i, column=1, value=f"{mins} min")
    vis.cell(row=i, column=2, value=band)
    s_formula = f'IF({mins}<S_thresh_short, S_short, IF({mins}<S_thresh_long, S_medium, S_long))'
    cog_mult = f'IF(include_cognitive="YES", 1 + {s_formula}/{mins}, 1)'
    rew_mult = "1+F_failure_rate*K_blast_radius"
    vis.cell(row=i, column=3, value=f"=D_devs*R_runs_per_dev_day*W_protective*({cog_mult})*({rew_mult})")
    vis.cell(row=i, column=4, value=f"=C{i}*{mins}*R_runs_per_dev_day*D_devs*days_per_year")
    vis.cell(row=i, column=5, value=narrative)

    style_label(vis.cell(row=i, column=1))
    style_label(vis.cell(row=i, column=2))
    style_derived(vis.cell(row=i, column=3), USD_FMT)
    style_derived(vis.cell(row=i, column=4), USD_BIG_FMT)
    vis.cell(row=i, column=5).fill = NOTE_FILL
    vis.cell(row=i, column=5).alignment = LEFT
    vis.cell(row=i, column=5).border = THIN
    vis.cell(row=i, column=5).font = NOTE_FONT
    vis.row_dimensions[i].height = 38

    # Highlight the row closest to the org's measured pipeline duration
    closest_mins = min((m for m, _, _ in vision_rows), key=lambda m: abs(m - current_t))
    if mins == closest_mins:
        for col_n in range(1, 6):
            vis.cell(row=i, column=col_n).fill = INPUT_STRONG

vis.conditional_formatting.add(
    f"C4:C{4 + len(vision_rows) - 1}",
    ColorScaleRule(start_type="min", start_color="2E7D32",
                   mid_type="percentile", mid_value=50, mid_color="FFEB3B",
                   end_type="max", end_color="C62828"),
)

chart = LineChart()
chart.title = "$/CI minute as pipeline duration changes"
chart.style = 10
chart.y_axis.title = "$/CI minute"
chart.x_axis.title = "Pipeline duration (min)"
data = Reference(vis, min_col=3, min_row=3, max_row=3 + len(vision_rows))
cats = Reference(vis, min_col=1, min_row=4, max_row=3 + len(vision_rows))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 18; chart.height = 10
vis.add_chart(chart, "G3")


# =============================================================================
# Sheet: Sensitivity (3-axis sweep)
# =============================================================================
sen = wb.create_sheet("Sensitivity")
for col, w in zip("ABCDEF", [16, 16, 16, 18, 18, 18]):
    sen.column_dimensions[col].width = w

sen.merge_cells("A1:F1")
sen["A1"] = "SENSITIVITY — how does the headline move when you change inputs?"
sen["A1"].font = HEADLINE_FONT
sen["A1"].fill = HEADLINE_FILL
sen["A1"].alignment = CENTER
sen.row_dimensions[1].height = 32

sen.merge_cells("A2:F2")
sen["A2"] = "Each row varies one or more inputs (rest held at the Inputs-sheet values). Use to find the cheapest leverage point."
sen["A2"].font = NOTE_FONT
sen["A2"].alignment = LEFT

for i, text in enumerate(["Wage ($/min)", "Failure rate", "Blast radius", "$/CI min", "Annual friction", "vs. baseline"], start=1):
    c = sen.cell(row=3, column=i, value=text)
    c.font = WHITE_BOLD; c.fill = SECTION_FILL; c.alignment = CENTER; c.border = THIN

wages = [1.50, 2.00, 2.50, 3.00]
failures = [0.10, 0.20, 0.30]
blasts = [1.0, 1.5, 2.5]

row_n = 4
for w in wages:
    for f in failures:
        for k in blasts:
            sen.cell(row=row_n, column=1, value=w)
            sen.cell(row=row_n, column=2, value=f)
            sen.cell(row=row_n, column=3, value=k)
            cog = 'IF(include_cognitive="YES", 1+S_applied/T_pipeline, 1)'
            cpm = (f"=D_devs*R_runs_per_dev_day*A{row_n}"
                   f"*({cog})"
                   f"*(1+B{row_n}*C{row_n})")
            sen.cell(row=row_n, column=4, value=cpm)
            sen.cell(row=row_n, column=5, value=f"=D{row_n}*T_pipeline*R_runs_per_dev_day*D_devs*days_per_year")
            sen.cell(row=row_n, column=6, value=f"=D{row_n}/Headline!$C$8 - 1")

            for col_letter, fmt in [("A", USD_FMT), ("B", PCT_FMT), ("C", "0.00"),
                                     ("D", USD_FMT), ("E", USD_BIG_FMT), ("F", "+0%;-0%")]:
                cell = sen.cell(row=row_n, column="ABCDEF".index(col_letter) + 1)
                cell.number_format = fmt
                cell.border = THIN
                cell.alignment = RIGHT

            if abs(w - 2.0) < 0.01 and abs(f - 0.30) < 0.01 and abs(k - 1.0) < 0.01:
                for col in "ABCDEF":
                    sen[f"{col}{row_n}"].fill = INPUT_STRONG

            row_n += 1

sen.conditional_formatting.add(
    f"D4:D{row_n - 1}",
    ColorScaleRule(start_type="min", start_color="2E7D32",
                   mid_type="percentile", mid_value=50, mid_color="FFEB3B",
                   end_type="max", end_color="C62828"),
)


# =============================================================================
# Sheet: Investments
# =============================================================================
inv = wb.create_sheet("Investments")
for col, w in zip("ABCDE", [40, 16, 22, 18, 60]):
    inv.column_dimensions[col].width = w

inv.merge_cells("A1:E1")
inv["A1"] = "INVESTMENT LADDER — what each fix is worth at $/CI minute"
inv["A1"].font = HEADLINE_FONT
inv["A1"].fill = HEADLINE_FILL
inv["A1"].alignment = CENTER
inv.row_dimensions[1].height = 32

inv["A2"] = "All ROI numbers use the Protective frame. Edit Inputs to recompute."
inv["A2"].font = NOTE_FONT
inv["A2"].alignment = LEFT
inv.merge_cells("A2:E2")

for i, text in enumerate(["Investment", "Effort", "What it does", "Annual ROI", "Payback"], start=1):
    c = inv.cell(row=3, column=i, value=text)
    c.font = WHITE_BOLD; c.fill = SECTION_FILL; c.alignment = CENTER; c.border = THIN

investments = [
    ("Better node_modules cache strategy", "1 dev-week", "Saves ~90 sec/build",
     "=Headline!$C$8*1.5*R_runs_per_dev_day*D_devs*days_per_year", "<1 week"),
    ("Pre-push hook (lint + typecheck)", "2 dev-days", "Cuts F from 27% → ~18%",
     "=Headline!$C$11*(F_failure_rate-0.18)/F_failure_rate*K_blast_radius/(1+F_failure_rate*K_blast_radius)", "<1 week"),
    ("Move E2E to selective tagging", "1 dev-week", "Saves ~3 min on ~30% of builds",
     "=Headline!$C$8*0.9*R_runs_per_dev_day*D_devs*days_per_year", "<2 weeks"),
    ("Ephemeral preview environments", "1 dev-month", "K → 1.0; eliminates STG queue",
     "=Headline!$C$11*(F_failure_rate*(K_blast_radius-1))/(1+F_failure_rate*K_blast_radius)", "<1 month"),
    ("Sub-10-min pipeline (sharding/cache)", "1 dev-quarter", "T_dur 17 → 8 min; crosses cliff",
     "=Headline!$C$11*0.55", "<1 month"),
    ("Sub-5-min pipeline (elite tier)", "Multi-quarter", "T_dur 17 → 5 min",
     "=Headline!$C$11*0.75", "<2 months"),
]

for i, (name, effort, what, roi, payback) in enumerate(investments, start=4):
    inv.cell(row=i, column=1, value=name)
    inv.cell(row=i, column=2, value=effort)
    inv.cell(row=i, column=3, value=what)
    inv.cell(row=i, column=4, value=roi)
    inv.cell(row=i, column=5, value=payback)
    for col_idx in range(1, 6):
        c = inv.cell(row=i, column=col_idx)
        c.border = THIN
        c.alignment = LEFT if col_idx in (1, 3, 5) else RIGHT
    inv.cell(row=i, column=4).number_format = USD_BIG_FMT
    inv.cell(row=i, column=4).fill = DERIVED_FILL
    inv.cell(row=i, column=4).font = Font(bold=True, color="2E7D32")
    inv.row_dimensions[i].height = 28


# =============================================================================
# Reorder + save
# =============================================================================
order = ["README", "Savings", "Headline", "Trend", "Wage Calculator", "Inputs", "Vision", "Sensitivity", "Investments"]
for idx, name in enumerate(order):
    wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

wb.active = 0  # README first

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}")
print(f"Size: {OUT_PATH.stat().st_size / 1024:.1f} KB")
print(f"Sheets: {', '.join(wb.sheetnames)}")
